"""
Unit Tests for Excel Parser Service

Tests the ExcelParser class which parses .xlsx files into structured data.
"""

import builtins
import io
import zipfile
from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
import pytest

from src.services.excel_parser import ExcelParser


class TestExcelParser:
    """Test suite for ExcelParser class."""

    def test_parse_simple_2_column_table(self, tmp_path: Path) -> None:
        """
        Test parsing a simple 2-column Excel table.

        Given: An Excel file with 2 columns (Name, Age) and 3 rows
        When: The file is parsed
        Then: Returns list of 3 dictionaries with Name and Age keys
        """
        # Create test Excel file
        file_path = tmp_path / "test_simple.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"

        # Add headers
        sheet["A1"] = "Name"
        sheet["B1"] = "Age"

        # Add data
        sheet["A2"] = "Alice"
        sheet["B2"] = 30
        sheet["A3"] = "Bob"
        sheet["B3"] = 25
        sheet["A4"] = "Charlie"
        sheet["B4"] = 35

        workbook.save(file_path)
        workbook.close()

        # Parse the file
        parser = ExcelParser()
        result = parser.parse_excel(file_path)

        # Assertions
        assert len(result) == 3
        assert result == [
            {"Name": "Alice", "Age": 30},
            {"Name": "Bob", "Age": 25},
            {"Name": "Charlie", "Age": 35},
        ]

    def test_parse_10_column_table(self, tmp_path: Path) -> None:
        """
        Test parsing a 10-column Excel table.

        Given: An Excel file with 10 columns and multiple rows
        When: The file is parsed
        Then: Returns list of dictionaries with all 10 columns
        """
        # Create test Excel file
        file_path = tmp_path / "test_10_columns.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Add headers for 10 columns
        headers = [f"Column{i}" for i in range(1, 11)]
        for col_idx, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_idx, value=header)

        # Add data row
        for col_idx, header in enumerate(headers, start=1):
            sheet.cell(row=2, column=col_idx, value=f"Data{col_idx}")

        workbook.save(file_path)
        workbook.close()

        # Parse the file
        parser = ExcelParser()
        result = parser.parse_excel(file_path)

        # Assertions
        assert len(result) == 1
        row_data = result[0]
        assert len(row_data) == 10
        for i in range(1, 11):
            assert f"Column{i}" in row_data
            assert row_data[f"Column{i}"] == f"Data{i}"

    def test_parse_handles_empty_cells(self, tmp_path: Path) -> None:
        """
        Test parsing Excel file with empty cells.

        Given: An Excel file with some empty cells
        When: The file is parsed
        Then: Empty cells are converted to empty strings
        """
        # Create test Excel file
        file_path = tmp_path / "test_empty_cells.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Add headers
        sheet["A1"] = "Name"
        sheet["B1"] = "Email"
        sheet["C1"] = "Phone"

        # Add data with empty cells
        sheet["A2"] = "Alice"
        sheet["B2"] = ""  # Empty email
        sheet["C2"] = "123-456-7890"

        sheet["A3"] = "Bob"
        sheet["B3"] = "bob@example.com"
        sheet["C3"] = None  # Empty phone

        workbook.save(file_path)
        workbook.close()

        # Parse the file
        parser = ExcelParser()
        result = parser.parse_excel(file_path)

        # Assertions
        assert len(result) == 2
        assert result[0] == {
            "Name": "Alice",
            "Email": "",
            "Phone": "123-456-7890",
        }
        assert result[1] == {
            "Name": "Bob",
            "Email": "bob@example.com",
            "Phone": "",
        }

    def test_parse_handles_different_data_types(self, tmp_path: Path) -> None:
        """
        Test parsing Excel file with different data types.

        Given: An Excel file with string, number, date, and boolean values
        When: The file is parsed
        Then: Data types are preserved appropriately (int, float, str, bool)
        """
        from datetime import date

        # Create test Excel file
        file_path = tmp_path / "test_data_types.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Add headers
        sheet["A1"] = "Name"
        sheet["B1"] = "Age"
        sheet["C1"] = "Salary"
        sheet["D1"] = "Active"
        sheet["E1"] = "BirthDate"

        # Add data with different types
        sheet["A2"] = "Alice"
        sheet["B2"] = 30  # Integer
        sheet["C2"] = 50000.50  # Float
        sheet["D2"] = True  # Boolean
        sheet["E2"] = date(1990, 5, 15)  # Date

        workbook.save(file_path)
        workbook.close()

        # Parse the file
        parser = ExcelParser()
        result = parser.parse_excel(file_path)

        # Assertions
        assert len(result) == 1
        row = result[0]
        assert row["Name"] == "Alice"
        assert row["Age"] == 30
        assert row["Salary"] == 50000.50
        assert row["Active"] is True
        # Date is returned as datetime object
        assert isinstance(row["BirthDate"], date)

    def test_parse_skips_empty_rows_by_default(self, tmp_path: Path) -> None:
        """
        Test that empty rows are skipped by default.

        Given: An Excel file with an empty row between data rows
        When: The file is parsed
        Then: Empty row is not included in results
        """
        # Create test Excel file
        file_path = tmp_path / "test_empty_rows.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Add headers
        sheet["A1"] = "Name"
        sheet["B1"] = "Value"

        # Add data with empty row
        sheet["A2"] = "Row1"
        sheet["B2"] = 100

        # Row 3 is empty (no values)

        sheet["A4"] = "Row2"
        sheet["B4"] = 200

        workbook.save(file_path)
        workbook.close()

        # Parse the file
        parser = ExcelParser()
        result = parser.parse_excel(file_path)

        # Assertions - empty row should be skipped
        assert len(result) == 2
        assert result == [
            {"Name": "Row1", "Value": 100},
            {"Name": "Row2", "Value": 200},
        ]

    def test_parse_includes_empty_rows_when_skip_false(self, tmp_path: Path) -> None:
        """
        Test that empty rows are included when skip_empty_rows=False.

        Given: An Excel file with an empty row
        When: The file is parsed with skip_empty_rows=False
        Then: Empty row is included with empty values
        """
        # Create test Excel file
        file_path = tmp_path / "test_include_empty.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Add headers
        sheet["A1"] = "Name"
        sheet["B1"] = "Value"

        # Add data
        sheet["A2"] = "Row1"
        sheet["B2"] = 100

        # Empty row
        sheet["A3"] = ""
        sheet["B3"] = ""

        workbook.save(file_path)
        workbook.close()

        # Parse with skip_empty_rows=False
        parser = ExcelParser()
        result = parser.parse_excel(file_path, skip_empty_rows=False)

        # Assertions - empty row should be included
        assert len(result) == 2
        assert result[0] == {"Name": "Row1", "Value": 100}
        # Empty row is included because skip_empty_rows=False
        # but the row still has values (empty strings)

    def test_parse_custom_sheet_name(self, tmp_path: Path) -> None:
        """
        Test parsing a specific sheet by name.

        Given: An Excel file with multiple sheets
        When: A specific sheet name is requested
        Then: Only that sheet is parsed
        """
        # Create test Excel file with multiple sheets
        file_path = tmp_path / "test_multi_sheet.xlsx"
        workbook = openpyxl.Workbook()
        sheet1 = workbook.active
        sheet1.title = "Data"

        # Add data to first sheet
        sheet1["A1"] = "Name"
        sheet1["A2"] = "From Data Sheet"

        # Create second sheet
        sheet2 = workbook.create_sheet("Summary")
        sheet2["A1"] = "Name"
        sheet2["A2"] = "From Summary Sheet"

        workbook.save(file_path)
        workbook.close()

        # Parse specific sheet
        parser = ExcelParser()
        result = parser.parse_excel(file_path, sheet_name="Summary")

        # Assertions
        assert len(result) == 1
        assert result[0] == {"Name": "From Summary Sheet"}

    def test_parse_default_sheet_when_not_specified(self, tmp_path: Path) -> None:
        """
        Test parsing uses active sheet when no sheet name specified.

        Given: An Excel file with multiple sheets
        When: No sheet name is specified
        Then: The active (first) sheet is parsed
        """
        # Create test Excel file
        file_path = tmp_path / "test_default_sheet.xlsx"
        workbook = openpyxl.Workbook()
        sheet1 = workbook.active
        sheet1.title = "First"

        sheet1["A1"] = "Name"
        sheet1["A2"] = "First Sheet Data"

        workbook.create_sheet("Second")
        workbook.save(file_path)
        workbook.close()

        # Parse without specifying sheet
        parser = ExcelParser()
        result = parser.parse_excel(file_path)

        # Assertions
        assert len(result) == 1
        assert result[0] == {"Name": "First Sheet Data"}

    def test_parse_file_not_found(self) -> None:
        """
        Test parsing a non-existent file.

        Given: A file path that doesn't exist
        When: Attempting to parse the file
        Then: FileNotFoundError is raised
        """
        parser = ExcelParser()
        with pytest.raises(FileNotFoundError, match="File not found"):
            parser.parse_excel("/nonexistent/path/file.xlsx")

    def test_parse_invalid_file_extension(self, tmp_path: Path) -> None:
        """
        Test parsing a file with wrong extension.

        Given: A file with .csv extension instead of .xlsx
        When: Attempting to parse the file as Excel
        Then: ValueError is raised
        """
        # Create a .csv file
        file_path = tmp_path / "test.csv"
        file_path.write_text("Name,Age\nAlice,30")

        parser = ExcelParser()
        with pytest.raises(ValueError, match="Invalid file type"):
            parser.parse_excel(file_path)

    def test_parse_invalid_excel_file(self, tmp_path: Path) -> None:
        """
        Test parsing an invalid Excel file.

        Given: A file with .xlsx extension but invalid content
        When: Attempting to parse the file
        Then: InvalidFileException or ValueError is raised
        """
        # Create invalid .xlsx file (just a text file with wrong extension)
        file_path = tmp_path / "invalid.xlsx"
        file_path.write_text("This is not a valid Excel file")

        parser = ExcelParser()
        with pytest.raises((InvalidFileException, ValueError)):
            parser.parse_excel(file_path)

    def test_parse_sheet_not_found(self, tmp_path: Path) -> None:
        """
        Test parsing a non-existent sheet.

        Given: An Excel file with a sheet named "Data"
        When: Attempting to parse a sheet named "NonExistent"
        Then: ValueError is raised with available sheet names
        """
        # Create test Excel file
        file_path = tmp_path / "test_sheet_not_found.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Data"
        sheet["A1"] = "Name"
        sheet["A2"] = "Test"
        workbook.save(file_path)
        workbook.close()

        # Try to parse non-existent sheet
        parser = ExcelParser()
        with pytest.raises(ValueError, match="Sheet 'NonExistent' not found"):
            parser.parse_excel(file_path, sheet_name="NonExistent")

    def test_parse_empty_sheet(self, tmp_path: Path) -> None:
        """
        Test parsing a sheet with no headers.

        Given: An Excel sheet with an empty first row
        When: Attempting to parse the sheet
        Then: ValueError is raised
        """
        # Create test Excel file with empty first row
        file_path = tmp_path / "test_empty_sheet.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        # Don't add any data - sheet is empty
        workbook.save(file_path)
        workbook.close()

        # Try to parse empty sheet
        parser = ExcelParser()
        with pytest.raises(ValueError, match="Sheet has no headers"):
            parser.parse_excel(file_path)

    def test_parse_skips_columns_without_headers(self, tmp_path: Path) -> None:
        """
        Test that columns without headers are skipped.

        Given: An Excel file where some columns have empty headers
        When: The file is parsed
        Then: Columns with empty headers are not included in results
        """
        # Create test Excel file
        file_path = tmp_path / "test_empty_headers.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Add headers with some empty
        sheet["A1"] = "Name"
        sheet["B1"] = ""  # Empty header
        sheet["C1"] = "Age"

        # Add data
        sheet["A2"] = "Alice"
        sheet["B2"] = "Ignored"
        sheet["C2"] = 30

        workbook.save(file_path)
        workbook.close()

        # Parse
        parser = ExcelParser()
        result = parser.parse_excel(file_path)

        # Assertions - column B should be ignored
        assert len(result) == 1
        assert result[0] == {"Name": "Alice", "Age": 30}

    def test_get_sheet_names(self, tmp_path: Path) -> None:
        """
        Test getting list of sheet names.

        Given: An Excel file with multiple sheets
        When: Getting sheet names
        Then: Returns list of all sheet names
        """
        # Create test Excel file
        file_path = tmp_path / "test_get_sheets.xlsx"
        workbook = openpyxl.Workbook()
        workbook.active.title = "First"
        workbook.create_sheet("Second")
        workbook.create_sheet("Third")
        workbook.save(file_path)
        workbook.close()

        # Get sheet names
        parser = ExcelParser()
        sheet_names = parser.get_sheet_names(file_path)

        # Assertions
        assert sheet_names == ["First", "Second", "Third"]

    def test_get_sheet_names_file_not_found(self) -> None:
        """
        Test getting sheet names from non-existent file.

        Given: A file path that doesn't exist
        When: Attempting to get sheet names
        Then: FileNotFoundError is raised
        """
        parser = ExcelParser()
        with pytest.raises(FileNotFoundError, match="File not found"):
            parser.get_sheet_names("/nonexistent/file.xlsx")

    def test_get_sheet_names_invalid_file(self, tmp_path: Path) -> None:
        """
        Test getting sheet names from invalid Excel file.

        Given: A file with .xlsx extension but invalid content
        When: Attempting to get sheet names
        Then: InvalidFileException is raised
        """
        # Create invalid .xlsx file
        file_path = tmp_path / "invalid.xlsx"
        file_path.write_text("Not a valid Excel file")

        parser = ExcelParser()
        with pytest.raises((InvalidFileException, ValueError)):
            parser.get_sheet_names(file_path)
