"""
Unit tests for Excel Template Filler service.

Tests Excel-to-Excel filling functionality for warehouse/delivery note generation.
"""

import pytest
from pathlib import Path
from unittest.mock import patch, MagicMock

from src.services.excel_template_filler import ExcelTemplateFiller, ExcelTemplateFillerError
from src.models.mapping import Mapping
from openpyxl import load_workbook


class TestExcelTemplateFillerCreation:
    """Test ExcelTemplateFiller initialization."""

    def test_create_filler(self):
        """Should create filler instance."""
        filler = ExcelTemplateFiller()
        assert filler is not None


class TestExcelTemplateFillerValidation:
    """Test input validation."""

    def test_fill_with_nonexistent_template_raises_error(self):
        """Should raise error if template file doesn't exist."""
        filler = ExcelTemplateFiller()
        
        with pytest.raises(FileNotFoundError):
            filler.fill_excel_template(
                template_path=Path("/nonexistent/template.xlsx"),
                data_row={"name": "test"},
                cell_mapping={"name": "A1"}
            )

    def test_fill_with_invalid_extension_raises_error(self):
        """Should raise error if file is not xlsx."""
        filler = ExcelTemplateFiller()
        
        with pytest.raises(ExcelTemplateFillerError):
            filler.fill_excel_template(
                template_path=Path("/tmp/template.docx"),
                data_row={"name": "test"},
                cell_mapping={"name": "A1"}
            )


class TestExcelTemplateFillerCellMapping:
    """Test cell mapping and data filling."""

    def test_map_cell_reference(self):
        """Should map placeholder to cell reference."""
        filler = ExcelTemplateFiller()
        
        # Cell mapping like "A1", "B2"
        mapping = {
            "订单号": "B2",
            "日期": "B3",
            "客户": "B4",
            "数量": "D2"
        }
        
        assert filler._get_cell_for_placeholder("订单号", mapping) == "B2"
        assert filler._get_cell_for_placeholder("日期", mapping) == "B3"
        assert filler._get_cell_for_placeholder("不存在", mapping) is None

    def test_fill_cell_value(self):
        """Should fill cell with correct value."""
        filler = ExcelTemplateFiller()
        
        data_row = {
            "订单号": "ORD-2024-001",
            "日期": "2024-02-15",
            "数量": 100
        }
        
        # Test value extraction
        assert filler._get_cell_value("订单号", data_row) == "ORD-2024-001"
        assert filler._get_cell_value("数量", data_row) == 100
        assert filler._get_cell_value("不存在", data_row) is None


class TestExcelTemplateMappingConfig:
    """Test Excel-specific mapping configuration."""

    def test_parse_cell_mapping_string(self):
        """Should parse cell mapping from string format."""
        filler = ExcelTemplateFiller()
        
        # Format: "placeholder:cell,placeholder:cell"
        mapping_str = "订单号:B2,日期:B3,客户:B4"
        
        result = filler._parse_cell_mapping(mapping_str)
        
        assert result == {
            "订单号": "B2",
            "日期": "B3",
            "客户": "B4"
        }

    def test_validate_cell_reference(self):
        """Should validate cell references."""
        filler = ExcelTemplateFiller()
        
        # Valid references
        assert filler._is_valid_cell_ref("A1") is True
        assert filler._is_valid_cell_ref("B10") is True
        assert filler._is_valid_cell_ref("AA100") is True
        
        # Invalid references
        assert filler._is_valid_cell_ref("1A") is False
        assert filler._is_valid_cell_ref("A") is False
        assert filler._is_valid_cell_ref("A1B") is False


class TestExcelTemplateBatchProcessing:
    """Test batch processing for multiple records."""

    def test_batch_fill_creates_multiple_sheets(self):
        """Should create one sheet per data row."""
        filler = ExcelTemplateFiller()
        
        data_rows = [
            {"订单号": "001", "数量": 10},
            {"订单号": "002", "数量": 20},
            {"订单号": "003", "数量": 30}
        ]
        
        # Create a real temporary file
        import tempfile
        import os
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            # Create a simple workbook
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Template"
            wb.save(tmp.name)
            tmp_path = tmp.name
        
        try:
            # Should create 3 sheets (one per row)
            result = filler.fill_excel_template_batch(
                template_path=Path(tmp_path),
                data_rows=data_rows,
                cell_mapping={"订单号": "B2", "数量": "B3"}
            )
            
            # Verify result is bytes
            assert isinstance(result, bytes)
            assert len(result) > 0
            
            # Load and verify
            from openpyxl import load_workbook
            import io
            result_wb = load_workbook(io.BytesIO(result))
            
            # Should have 3 sheets
            assert len(result_wb.sheetnames) == 3
            
        finally:
            os.unlink(tmp_path)

    def test_batch_fill_separate_files(self):
        """Option: create separate file per row."""
        filler = ExcelTemplateFiller()
        
        data_rows = [
            {"订单号": "001", "数量": 10},
            {"订单号": "002", "数量": 20}
        ]
        
        # Create a real temporary file
        import tempfile
        import os
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            from openpyxl import Workbook
            wb = Workbook()
            wb.save(tmp.name)
            tmp_path = tmp.name
        
        try:
            results = filler.fill_excel_template_batch_separate(
                template_path=Path(tmp_path),
                data_rows=data_rows,
                cell_mapping={"订单号": "B2", "数量": "B3"},
                output_prefix="出库单"
            )
            
            # Should return list of (filename, bytes) tuples
            assert len(results) == 2
            assert results[0][0].startswith("出库单")
            assert results[0][0].endswith(".xlsx")
            assert isinstance(results[0][1], bytes)
            
        finally:
            os.unlink(tmp_path)


class TestWarehouseDeliveryNoteScenario:
    """Test the specific warehouse delivery note use case."""

    def test_delivery_note_template_structure(self):
        """Typical delivery note template structure."""
        filler = ExcelTemplateFiller()
        
        # Typical warehouse delivery note has:
        # - 单号、日期、客户、产品、数量、仓库、经手人
        template_cells = {
            "出库单号": "B2",
            "日期": "B3", 
            "客户名称": "B4",
            "产品编号": "B5",
            "产品名称": "B6",
            "规格": "B7",
            "数量": "D2",
            "单位": "D3",
            "仓库": "D4",
            "经手人": "D5"
        }
        
        production_data = {
            "出库单号": "CK-20240215-001",
            "日期": "2024-02-15",
            "客户名称": "ABC有限公司",
            "产品编号": "PROD-001",
            "产品名称": "工业配件A",
            "规格": "标准型",
            "数量": 1000,
            "单位": "件",
            "仓库": "A1仓库",
            "经手人": "张三"
        }
        
        # Verify mapping logic
        for key, cell in template_cells.items():
            value = production_data.get(key)
            assert value is not None, f"Missing data for {key}"

    def test_formula_preservation(self):
        """Excel formulas should be preserved in template."""
        filler = ExcelTemplateFiller()
        
        import tempfile
        import os
        from openpyxl import Workbook
        import io
        
        # Create template with formula
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = Workbook()
            ws = wb.active
            ws["B2"] = 10  # 单价
            ws["C2"] = 5   # 数量
            ws["D2"] = "=B2*C2"  # 总价公式
            wb.save(tmp.name)
            tmp_path = tmp.name
        
        try:
            # Fill template (modifying B2 and C2)
            result = filler.fill_excel_template(
                template_path=Path(tmp_path),
                data_row={"单价": 20, "数量": 3},
                cell_mapping={"单价": "B2", "数量": "C2"}
            )
            
            # Verify result
            assert isinstance(result, bytes)
            
            # Load and check formula preserved
            result_wb = load_workbook(io.BytesIO(result))
            result_ws = result_wb.active
            
            # Values should be updated
            assert result_ws["B2"].value == 20
            assert result_ws["C2"].value == 3
            
            # Formula should be preserved in D2
            # Note: openpyxl stores formulas as strings starting with =
            assert "=B2*C2" in str(result_ws["D2"].value)
            
        finally:
            os.unlink(tmp_path)
