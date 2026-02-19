"""
Fill Application - Excel Parser Service

Parses Excel (.xlsx) files into structured data (list of dictionaries).
Uses openpyxl library for reading Excel files.
"""

from datetime import date, datetime
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException


class ExcelParser:
    """
    Parser for Excel (.xlsx) files.

    Extracts data from Excel sheets and converts them to a list of dictionaries,
    where each dictionary represents a row with column headers as keys.
    """

    DEFAULT_SHEET_NAME = "Sheet1"

    @staticmethod
    def parse_excel(
        file_path: str | Path,
        sheet_name: str | None = None,
        skip_empty_rows: bool = True,
    ) -> list[dict[str, Any]]:
        """
        Parse an Excel file and return structured data.

        Args:
            file_path: Path to the Excel file (.xlsx)
            sheet_name: Name of the sheet to parse. If None, uses the first sheet.
            skip_empty_rows: If True, skips rows where all values are empty

        Returns:
            List of dictionaries, where each dictionary represents a row with
            column headers as keys and cell values as values.

        Raises:
            FileNotFoundError: If the file doesn't exist
            InvalidFileException: If the file is not a valid Excel file
            ValueError: If the sheet is not found or has no headers

        Examples:
            >>> parser = ExcelParser()
            >>> data = parser.parse_excel("data.xlsx")
            >>> # Returns: [{"Name": "John", "Age": 30}, ...]
        """
        file_path = Path(file_path)

        # Validate file exists
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        # Validate file extension
        if file_path.suffix.lower() != ".xlsx":
            raise ValueError(f"Invalid file type. Expected .xlsx, got: {file_path.suffix}")

        # Load workbook
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        except InvalidFileException as e:
            raise InvalidFileException(f"Invalid Excel file: {e}") from e
        except Exception as e:
            raise ValueError(f"Error loading Excel file: {e}") from e

        # Select sheet
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(
                    f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(workbook.sheetnames)}"
                )
            sheet = workbook[sheet_name]
        else:
            # Use first sheet
            sheet = workbook.active
            if sheet is None:
                raise ValueError("Workbook has no sheets")

        # Extract headers from first row with error handling
        headers: list[str] = []
        try:
            first_row = sheet[1]
        except IndexError:
            # Sheet exists but has no rows
            raise ValueError("Sheet has no headers (first row is empty)")

        for cell in first_row:
            value = cell.value
            if value is None:
                headers.append("")  # Empty header
            elif isinstance(value, str):
                headers.append(value.strip())
            else:
                headers.append(str(value))

        # Validate headers
        if not any(headers):
            raise ValueError("Sheet has no headers (first row is empty)")

        # Extract data rows
        data: list[dict[str, Any]] = []

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # Create row dictionary
            row_data: dict[str, Any] = {}

            for col_idx, (header, value) in enumerate(zip(headers, row), start=1):
                if not header:
                    # Skip columns without headers
                    continue

                # Handle different cell value types
                if value is None:
                    processed_value = ""
                elif isinstance(value, (int, float, str, bool, datetime, date)):
                    # Preserve datetime and date objects
                    processed_value = value
                else:
                    # Convert other types to string
                    processed_value = str(value)

                row_data[header] = processed_value

            # Skip empty rows if requested
            if skip_empty_rows:
                # Check if all values are empty strings or None
                if all(v == "" or v is None for v in row_data.values()):
                    continue

            # Add row to data (even if empty, if skip_empty_rows is False)
            if row_data or not skip_empty_rows:
                data.append(row_data)

        workbook.close()
        return data

    @staticmethod
    def get_sheet_names(file_path: str | Path) -> list[str]:
        """
        Get list of sheet names in the Excel file.

        Args:
            file_path: Path to the Excel file

        Returns:
            List of sheet names

        Raises:
            FileNotFoundError: If the file doesn't exist
            InvalidFileException: If the file is not a valid Excel file
        """
        file_path = Path(file_path)

        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            sheet_names = workbook.sheetnames
            workbook.close()
            return sheet_names
        except InvalidFileException as e:
            raise InvalidFileException(f"Invalid Excel file: {e}") from e
        except BadZipFile as e:
            raise InvalidFileException(f"Invalid Excel file: {e}") from e
