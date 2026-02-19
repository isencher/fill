"""
Excel Template Filler Service

Fills Excel templates (.xlsx) with data from rows.
Supports warehouse delivery notes, production reports, and other Excel-to-Excel workflows.
"""

import io
from pathlib import Path
from typing import Any, Optional

try:
    import openpyxl
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExcelTemplateFillerError(Exception):
    """Custom exception for Excel template filling errors."""
    
    def __init__(self, message: str) -> None:
        self.message = message
        super().__init__(self.message)


class ExcelTemplateFiller:
    """
    Fills Excel templates with data values.
    
    Supports:
    - Cell-based placeholder replacement
    - Formula preservation
    - Batch processing (one sheet per row or separate files)
    
    Typical use cases:
    - Warehouse delivery notes (出库单)
    - Production reports (生产日报)
    - Invoice forms (发票)
    """
    
    def __init__(self) -> None:
        """Initialize the Excel template filler."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel template filling. "
                "Install it with: pip install openpyxl"
            )
    
    def fill_excel_template(
        self,
        template_path: str | Path,
        data_row: dict[str, Any],
        cell_mapping: dict[str, str],
        sheet_name: Optional[str] = None,
    ) -> bytes:
        """
        Fill an Excel template with a single data row.
        
        Args:
            template_path: Path to Excel template file (.xlsx)
            data_row: Dictionary of column names to data values
            cell_mapping: Mapping of placeholder names to cell references
                         e.g., {"订单号": "B2", "日期": "B3"}
            sheet_name: Target sheet name (default: active sheet)
            
        Returns:
            Filled Excel file as bytes
            
        Raises:
            FileNotFoundError: If template file doesn't exist
            ExcelTemplateFillerError: If template format is invalid
        """
        template_path = Path(template_path)
        
        # Validate extension first
        if template_path.suffix.lower() != ".xlsx":
            raise ExcelTemplateFillerError(
                f"Unsupported template format: {template_path.suffix}. "
                f"Only .xlsx files are supported."
            )
        
        # Then check file exists
        if not template_path.exists():
            raise FileNotFoundError(f"Template file not found: {template_path}")
        
        # Load workbook
        try:
            wb = load_workbook(str(template_path))
        except Exception as e:
            raise ExcelTemplateFillerError(f"Failed to load Excel template: {e}")
        
        # Get target worksheet
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise ExcelTemplateFillerError(f"Sheet not found: {sheet_name}")
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        # Fill cells according to mapping
        for placeholder, cell_ref in cell_mapping.items():
            if not self._is_valid_cell_ref(cell_ref):
                continue
            
            value = self._get_cell_value(placeholder, data_row)
            if value is not None:
                ws[cell_ref] = value
        
        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()
    
    def fill_excel_template_batch(
        self,
        template_path: str | Path,
        data_rows: list[dict[str, Any]],
        cell_mapping: dict[str, str],
    ) -> bytes:
        """
        Fill Excel template with multiple data rows.
        Creates one worksheet per data row.
        
        Args:
            template_path: Path to Excel template file
            data_rows: List of data dictionaries (one per row/record)
            cell_mapping: Mapping of placeholders to cell references
            
        Returns:
            Excel workbook with multiple sheets as bytes
            
        Example:
            >>> data_rows = [
            ...     {"订单号": "001", "数量": 100},
            ...     {"订单号": "002", "数量": 200}
            ... ]
            >>> result = filler.fill_excel_template_batch(
            ...     template_path="template.xlsx",
            ...     data_rows=data_rows,
            ...     cell_mapping={"订单号": "B2", "数量": "B3"}
            ... )
            # Result has 2 sheets, one per order
        """
        template_path = Path(template_path)
        
        if not template_path.exists():
            raise FileNotFoundError(f"Template file not found: {template_path}")
        
        if not data_rows:
            raise ExcelTemplateFillerError("No data rows provided")
        
        # Load workbook
        wb = load_workbook(str(template_path))
        original_sheet = wb.active
        original_title = original_sheet.title
        
        # Fill first sheet with first row
        for placeholder, cell_ref in cell_mapping.items():
            value = self._get_cell_value(placeholder, data_rows[0])
            if value is not None and self._is_valid_cell_ref(cell_ref):
                original_sheet[cell_ref] = value
        
        # For remaining rows, copy sheet and fill
        for i, data_row in enumerate(data_rows[1:], start=2):
            # Copy worksheet
            new_sheet = wb.copy_worksheet(original_sheet)
            new_sheet.title = f"{original_title}_{i}"
            
            # Fill with data
            for placeholder, cell_ref in cell_mapping.items():
                value = self._get_cell_value(placeholder, data_row)
                if value is not None and self._is_valid_cell_ref(cell_ref):
                    new_sheet[cell_ref] = value
        
        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()
    
    def fill_excel_template_batch_separate(
        self,
        template_path: str | Path,
        data_rows: list[dict[str, Any]],
        cell_mapping: dict[str, str],
        output_prefix: str = "output",
    ) -> list[tuple[str, bytes]]:
        """
        Create separate Excel file for each data row.
        
        Args:
            template_path: Path to Excel template file
            data_rows: List of data dictionaries
            cell_mapping: Mapping of placeholders to cell references
            output_prefix: Prefix for generated filenames
            
        Returns:
            List of (filename, file_bytes) tuples
            
        Example:
            >>> results = filler.fill_excel_template_batch_separate(
            ...     template_path="template.xlsx",
            ...     data_rows=[{"订单号": "001"}, {"订单号": "002"}],
            ...     cell_mapping={"订单号": "B2"},
            ...     output_prefix="出库单"
            ... )
            >>> results[0][0]  # "出库单_001.xlsx"
        """
        template_path = Path(template_path)
        results = []
        
        for i, data_row in enumerate(data_rows, start=1):
            # Fill template
            file_bytes = self.fill_excel_template(
                template_path=template_path,
                data_row=data_row,
                cell_mapping=cell_mapping
            )
            
            # Generate filename
            # Try to use identifier from data (e.g., order number)
            identifier = data_row.get("订单号") or data_row.get("单号") or str(i)
            filename = f"{output_prefix}_{identifier}.xlsx"
            
            results.append((filename, file_bytes))
        
        return results
    
    def _get_cell_for_placeholder(
        self,
        placeholder: str,
        cell_mapping: dict[str, str]
    ) -> Optional[str]:
        """Get cell reference for a placeholder."""
        return cell_mapping.get(placeholder)
    
    def _get_cell_value(
        self,
        placeholder: str,
        data_row: dict[str, Any]
    ) -> Any:
        """Get value from data row for a placeholder."""
        # Direct match
        if placeholder in data_row:
            return data_row[placeholder]
        
        # Try common variations
        variations = self._get_placeholder_variations(placeholder)
        for var in variations:
            if var in data_row:
                return data_row[var]
        
        return None
    
    def _get_placeholder_variations(self, placeholder: str) -> list[str]:
        """Get possible variations of a placeholder name."""
        variations = [placeholder]
        
        # Common suffixes/prefixes
        suffixes = ["名称", "编号", "号", "日期", "时间"]
        for suffix in suffixes:
            if placeholder.endswith(suffix):
                variations.append(placeholder[:-len(suffix)])
            else:
                variations.append(placeholder + suffix)
        
        return variations
    
    def _is_valid_cell_ref(self, cell_ref: str) -> bool:
        """
        Validate Excel cell reference.
        
        Valid: A1, B10, AA100
        Invalid: 1A, A, A1B
        """
        if not cell_ref or not isinstance(cell_ref, str):
            return False
        
        cell_ref = cell_ref.upper().strip()
        
        # Split into column letters and row numbers
        col = ""
        row = ""
        
        for char in cell_ref:
            if char.isalpha():
                if row:  # Letter after number - invalid
                    return False
                col += char
            elif char.isdigit():
                row += char
            else:
                return False
        
        # Must have both column and row
        if not col or not row:
            return False
        
        # Column must be A-Z+
        for c in col:
            if not ('A' <= c <= 'Z'):
                return False
        
        # Row must be positive integer
        try:
            row_num = int(row)
            if row_num < 1:
                return False
        except ValueError:
            return False
        
        return True
    
    def _parse_cell_mapping(self, mapping_str: str) -> dict[str, str]:
        """
        Parse cell mapping from string format.
        
        Format: "placeholder:cell,placeholder:cell"
        Example: "订单号:B2,日期:B3,数量:D2"
        """
        result = {}
        if not mapping_str:
            return result
        
        pairs = mapping_str.split(",")
        for pair in pairs:
            if ":" not in pair:
                continue
            placeholder, cell = pair.split(":", 1)
            placeholder = placeholder.strip()
            cell = cell.strip().upper()
            
            if self._is_valid_cell_ref(cell):
                result[placeholder] = cell
        
        return result


def fill_excel_template(
    template_path: str | Path,
    data_row: dict[str, Any],
    cell_mapping: dict[str, str],
) -> bytes:
    """
    Convenience function to fill an Excel template.
    
    Args:
        template_path: Path to Excel template file
        data_row: Data dictionary
        cell_mapping: Mapping of placeholders to cell references
        
    Returns:
        Filled Excel file as bytes
    """
    filler = ExcelTemplateFiller()
    return filler.fill_excel_template(template_path, data_row, cell_mapping)
